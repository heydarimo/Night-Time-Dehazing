import torch
import time
import argparse
import os
import math
import random
import numpy as np
from torch.utils.data import DataLoader
# from tensorboardX import SummaryWriter
import torch.nn.functional as F
from torchvision.models import vgg16
from torchvision.utils import save_image as imwrite
import logging
import torch.nn as nn
from model_flashinternimage_histoforme2 import fusion_net_histoformer2, Discriminator
from paired_dehaze_dataset import PairedDehazeDataset

from tqdm import tqdm
from pytorch_msssim import msssim, ssim
from perceptual import LossNetwork
import re
from torch.nn.parallel import DataParallel
from openpyxl import Workbook, load_workbook


parser = argparse.ArgumentParser(description='Histoformer Dehaze')
parser.add_argument('-learning_rate', help='Set the learning rate', default=1e-4, type=float)
parser.add_argument('-train_batch_size', help='Set the training batch size', default=2, type=int)
parser.add_argument('-train_epoch', help='Set the training epoch', default=5000, type=int)

# ---- universal paired loader arguments ----
parser.add_argument('--dataset_root', type=str, required=True,
                    help='Path to dataset root, e.g. /project/6050328/heydam3/paper/dataset/Dense-Haze')
parser.add_argument('--crop_size', type=int, default=384,
                    help='Random crop size for training')
parser.add_argument('--num_workers', type=int, default=4,
                    help='Number of dataloader workers')

parser.add_argument('--model_save_dir', type=str, default='Checkpoints')
parser.add_argument('--denet_save_dir', type=str, default='Checkpoints')
parser.add_argument('--log_dir', type=str, default=None)

# --- test / logging --- #
parser.add_argument('--output_dir', type=str, default='experiments/Histoformer/test_result')
parser.add_argument('--log_txt', type=str, default='train_loss_mixed_data.txt',
                    help='Training log txt filename (saved in current working directory)')
parser.add_argument('--excel_txt', type=str, default='test_metrics_live.xlsx',
                    help='Excel file for live test PSNR/SSIM logging')
parser.add_argument('-test_batch_size', help='Set the testing batch size', default=1, type=int)

# --- reproducible resume --- #
parser.add_argument('--resume_ckpt', type=str, default='',
                    help='Path to full resumable checkpoint (.pth)')
parser.add_argument('--seed', type=int, default=1234,
                    help='Random seed for reproducibility')

args = parser.parse_args()

# --- train --- #
learning_rate = args.learning_rate
train_batch_size = args.train_batch_size
train_epoch = args.train_epoch
dataset_root = args.dataset_root
crop_size = args.crop_size

# --- test --- #
test_batch_size = args.test_batch_size

# --- output folders --- #
os.makedirs(args.model_save_dir, exist_ok=True)
os.makedirs(args.denet_save_dir, exist_ok=True)
os.makedirs(args.output_dir, exist_ok=True)

# --- device --- #
device = 'cuda:0'
print(device)

# ----------------------------
# Reproducibility helpers
# ----------------------------
def set_all_seeds(seed: int):
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    torch.cuda.manual_seed_all(seed)


def seed_worker(worker_id):
    worker_seed = torch.initial_seed() % 2**32
    random.seed(worker_seed)
    np.random.seed(worker_seed)


def get_rng_state_dict():
    state = {
        'python_random_state': random.getstate(),
        'numpy_random_state': np.random.get_state(),
        'torch_random_state': torch.get_rng_state().cpu(),
    }
    if torch.cuda.is_available():
        state['torch_cuda_random_state_all'] = [s.cpu() for s in torch.cuda.get_rng_state_all()]
    else:
        state['torch_cuda_random_state_all'] = None
    return state


def set_rng_state_dict(state):
    random.setstate(state['python_random_state'])
    np.random.set_state(state['numpy_random_state'])

    cpu_rng = state['torch_random_state']
    if not isinstance(cpu_rng, torch.Tensor):
        cpu_rng = torch.tensor(cpu_rng, dtype=torch.uint8)
    cpu_rng = cpu_rng.detach().cpu().to(dtype=torch.uint8)
    torch.set_rng_state(cpu_rng)

    if torch.cuda.is_available() and state.get('torch_cuda_random_state_all') is not None:
        cuda_states = []
        for s in state['torch_cuda_random_state_all']:
            if not isinstance(s, torch.Tensor):
                s = torch.tensor(s, dtype=torch.uint8)
            cuda_states.append(s.detach().cpu().to(dtype=torch.uint8))
        torch.cuda.set_rng_state_all(cuda_states)


def atomic_torch_save(obj, path):
    tmp_path = path + '.tmp'
    torch.save(obj, tmp_path)
    os.replace(tmp_path, path)


def save_resume_checkpoint(path,
                           epoch,
                           iteration,
                           model,
                           dnet,
                           g_optimizer,
                           d_optimizer,
                           scheduler_g,
                           scheduler_d,
                           best_psnr,
                           best_ssim_at_best_psnr,
                           best_epoch,
                           train_loader_generator):
    state = {
        'epoch': epoch,
        'iteration': iteration,
        'model_state_dict': model.state_dict(),
        'dnet_state_dict': dnet.state_dict(),
        'g_optimizer_state_dict': g_optimizer.state_dict(),
        'd_optimizer_state_dict': d_optimizer.state_dict(),
        'scheduler_g_state_dict': scheduler_g.state_dict(),
        'scheduler_d_state_dict': scheduler_d.state_dict(),
        'best_psnr': best_psnr,
        'best_ssim_at_best_psnr': best_ssim_at_best_psnr,
        'best_epoch': best_epoch,
        'rng_state': get_rng_state_dict(),
        'train_loader_generator_state': train_loader_generator.get_state().cpu(),
    }
    atomic_torch_save(state, path)


def load_resume_checkpoint(path,
                           model,
                           dnet,
                           g_optimizer,
                           d_optimizer,
                           scheduler_g,
                           scheduler_d,
                           train_loader_generator,
                           device='cuda:0'):
    # IMPORTANT: load checkpoint on CPU so RNG states remain CPU ByteTensors
    ckpt = torch.load(path, map_location='cpu')

    model.load_state_dict(ckpt['model_state_dict'])
    dnet.load_state_dict(ckpt['dnet_state_dict'])
    g_optimizer.load_state_dict(ckpt['g_optimizer_state_dict'])
    d_optimizer.load_state_dict(ckpt['d_optimizer_state_dict'])
    scheduler_g.load_state_dict(ckpt['scheduler_g_state_dict'])
    scheduler_d.load_state_dict(ckpt['scheduler_d_state_dict'])

    set_rng_state_dict(ckpt['rng_state'])

    gen_state = ckpt['train_loader_generator_state']
    if not isinstance(gen_state, torch.Tensor):
        gen_state = torch.tensor(gen_state, dtype=torch.uint8)
    train_loader_generator.set_state(gen_state.detach().cpu().to(dtype=torch.uint8))

    start_epoch = ckpt['epoch'] + 1
    iteration = ckpt.get('iteration', 0)
    best_psnr = ckpt.get('best_psnr', -1e9)
    best_ssim_at_best_psnr = ckpt.get('best_ssim_at_best_psnr', -1e9)
    best_epoch = ckpt.get('best_epoch', -1)

    return start_epoch, iteration, best_psnr, best_ssim_at_best_psnr, best_epoch


set_all_seeds(args.seed)

# --- network --- #
MyEnsembleNet = fusion_net_histoformer2()
MyEnsembleNet = MyEnsembleNet.to(device)
print('MyEnsembleNet parameters:', sum(param.numel() for param in MyEnsembleNet.parameters()))

DNet = Discriminator()

# MyEnsembleNet = DataParallel(MyEnsembleNet, device_ids=[0, 1])
# DNet = DataParallel(DNet, device_ids=[0, 1])

# --- optimizer --- #
G_optimizer = torch.optim.Adam(MyEnsembleNet.parameters(), lr=learning_rate)
scheduler_G = torch.optim.lr_scheduler.MultiStepLR(G_optimizer, milestones=[2000, 3000, 4000], gamma=0.5)

D_optim = torch.optim.Adam(DNet.parameters(), lr=learning_rate)
scheduler_D = torch.optim.lr_scheduler.MultiStepLR(D_optim, milestones=[2500, 3000, 4000], gamma=0.5)

# --- datasets --- #
train_dataset = PairedDehazeDataset(
    dataset_root=dataset_root,
    split='train',
    crop_size=crop_size,
    augment_data=True
)

train_loader_generator = torch.Generator()
train_loader_generator.manual_seed(args.seed)

train_loader = DataLoader(
    dataset=train_dataset,
    batch_size=train_batch_size,
    shuffle=True,
    num_workers=args.num_workers,
    pin_memory=True,
    worker_init_fn=seed_worker,
    generator=train_loader_generator
)

test_dataset = PairedDehazeDataset(
    dataset_root=dataset_root,
    split='test',
    crop_size=None,
    augment_data=False
)

test_loader = DataLoader(
    dataset=test_dataset,
    batch_size=test_batch_size,
    shuffle=False,
    num_workers=args.num_workers,
    pin_memory=True,
    worker_init_fn=seed_worker
)

MyEnsembleNet = MyEnsembleNet.to(device)
DNet = DNet.to(device)

# --- perceptual loss --- #
vgg_model = vgg16(pretrained=True)
vgg_model = vgg_model.features[:16].to(device)
for param in vgg_model.parameters():
    param.requires_grad = False
loss_network = LossNetwork(vgg_model)
loss_network.eval()
msssim_loss = msssim


def tensor_psnr(pred, target):
    mse = torch.mean((pred - target) ** 2)
    if mse.item() == 0:
        return 100.0
    psnr = 10 * torch.log10(1.0 / mse)
    return psnr.item()


def tensor_ssim(pred, target):
    return ssim(pred, target, size_average=True).item()


def pad_to_multiple(x, multiple=32):
    _, _, h, w = x.shape
    new_h = math.ceil(h / multiple) * multiple
    new_w = math.ceil(w / multiple) * multiple

    pad_h = new_h - h
    pad_w = new_w - w

    if pad_h == 0 and pad_w == 0:
        return x, h, w

    x = F.pad(x, (0, pad_w, 0, pad_h), mode='reflect')
    return x, h, w


def crop_to_original(x, orig_h, orig_w):
    return x[:, :, :orig_h, :orig_w]


def evaluate_test_metrics(model, dataloader, device, pad_multiple=32):
    model.eval()
    total_psnr = 0.0
    total_ssim = 0.0
    count = 0

    with torch.no_grad():
        for hazy, clean in tqdm(dataloader, desc='Testing', leave=False):
            hazy = hazy.to(device)
            clean = clean.to(device)

            hazy_pad, orig_h, orig_w = pad_to_multiple(hazy, multiple=pad_multiple)

            output = model(hazy_pad)
            output = torch.clamp(output, 0.0, 1.0)
            output = crop_to_original(output, orig_h, orig_w)

            batch_size = output.shape[0]
            for b in range(batch_size):
                pred_b = output[b:b + 1]
                clean_b = clean[b:b + 1]
                total_psnr += tensor_psnr(pred_b, clean_b)
                total_ssim += tensor_ssim(pred_b, clean_b)
                count += 1

    avg_psnr = total_psnr / count
    avg_ssim = total_ssim / count
    return avg_psnr, avg_ssim


def init_excel_file(excel_path):
    if not os.path.exists(excel_path):
        wb = Workbook()
        ws = wb.active
        ws.title = 'test_metrics'
        ws.append(['epoch', 'avg_psnr', 'avg_ssim', 'is_best_psnr_so_far'])
        wb.save(excel_path)


def append_excel_row(excel_path, epoch, avg_psnr, avg_ssim, is_best):
    wb = load_workbook(excel_path)
    ws = wb['test_metrics']
    ws.append([epoch, avg_psnr, avg_ssim, 'YES' if is_best else 'NO'])
    wb.save(excel_path)


# --- initialize live Excel file --- #
init_excel_file(args.excel_txt)

# --- best checkpoint tracking: PSNR primary, SSIM tie-breaker --- #
best_psnr = -1e9
best_ssim_at_best_psnr = -1e9
best_epoch = -1

# --- resume state --- #
start_epoch = 0
iteration = 0

if args.resume_ckpt:
    if not os.path.isfile(args.resume_ckpt):
        raise FileNotFoundError(f'Resume checkpoint not found: {args.resume_ckpt}')
    start_epoch, iteration, best_psnr, best_ssim_at_best_psnr, best_epoch = load_resume_checkpoint(
        args.resume_ckpt,
        MyEnsembleNet,
        DNet,
        G_optimizer,
        D_optim,
        scheduler_G,
        scheduler_D,
        train_loader_generator,
        device=device
    )
    print(f'Resumed from {args.resume_ckpt}')
    print(f'  start_epoch = {start_epoch}')
    print(f'  iteration   = {iteration}')
    print(f'  best_epoch  = {best_epoch}')
    print(f'  best_psnr   = {best_psnr}')
    print(f'  best_ssim   = {best_ssim_at_best_psnr}')

# --- rolling resume checkpoint paths (overwritten, not accumulated) --- #
rolling_resume_path = os.path.join(args.model_save_dir, 'resume_latest.pth')
pre_eval_resume_path = os.path.join(args.model_save_dir, 'pre_eval_resume_latest.pth')

# --- training --- #
print('Start training')

for epoch in range(start_epoch, train_epoch):
    print('Epoch: ' + str(epoch))

    running_loss_l1 = 0
    running_loss_percep = 0
    running_loss_ssim = 0
    running_total_loss = 0

    start_time = time.time()
    MyEnsembleNet.train()
    DNet.train()
    pbar = tqdm(train_loader)

    for batch_idx, (hazy, clean) in enumerate(pbar):
        iteration += 1
        hazy = hazy.to(device)
        clean = clean.to(device)

        output = MyEnsembleNet(hazy)

        DNet.zero_grad()
        real_out = DNet(clean).mean()
        fake_out = DNet(output).mean()
        D_loss = 1 - real_out + fake_out
        D_loss.backward(retain_graph=True)

        adversarial_loss = torch.mean(1 - fake_out)

        MyEnsembleNet.zero_grad()
        adversarial_loss = torch.mean(1 - fake_out)
        smooth_loss_l1 = F.smooth_l1_loss(output, clean)
        perceptual_loss = loss_network(output, clean)
        msssim_loss_ = -msssim_loss(output, clean, normalize=True)
        total_loss = smooth_loss_l1 + 0.01 * perceptual_loss + 0.0005 * adversarial_loss + 0.2 * msssim_loss_
        total_loss.backward()

        pbar.set_postfix(loss=f"{total_loss.item():.4f}", lr=f"{G_optimizer.param_groups[0]['lr']:.2e}")

        D_optim.step()
        G_optimizer.step()

        running_total_loss += total_loss.item() * output.shape[0]
        running_loss_l1 += smooth_loss_l1.item() * output.shape[0]
        running_loss_percep += perceptual_loss.item() * output.shape[0]
        running_loss_ssim += msssim_loss_.item() * output.shape[0]

    # move scheduler step to end of epoch to avoid warning
    scheduler_G.step()
    scheduler_D.step()

    epoch_time = time.time() - start_time
    print(f"Epoch time: {epoch_time/60:.2f} min ({epoch_time:.1f} sec)")

    num_train_samples = len(train_dataset)
    avr_total_loss = running_total_loss / num_train_samples
    avr_l1_loss = running_loss_l1 / num_train_samples
    avr_percep_loss = running_loss_percep / num_train_samples
    avr_ssim_loss = running_loss_ssim / num_train_samples

    print('  avr_total_loss: ', avr_total_loss,
          '  avrl1_loss:  ', avr_l1_loss,
          '  avr_p_loss:  ', avr_percep_loss,
          '  avr_ssim_loss: ', avr_ssim_loss)

    with open(args.log_txt, 'a') as f:
        f.write('epoch:  ' + str(epoch) +
                '  lr_model:  ' + str(G_optimizer.param_groups[0]['lr']) +
                '  Dnet_lr:  ' + str(D_optim.param_groups[0]['lr']) +
                '  avr_total_loss: ' + str(avr_total_loss) +
                '  avr_l1_loss:  ' + str(avr_l1_loss) +
                '  avr_p_loss:  ' + str(avr_percep_loss) +
                '  avr_ssim_loss: ' + str(avr_ssim_loss) + '\n')

    # --- rolling epoch-end resume checkpoint (overwritten each epoch) --- #
    save_resume_checkpoint(
        rolling_resume_path,
        epoch,
        iteration,
        MyEnsembleNet,
        DNet,
        G_optimizer,
        D_optim,
        scheduler_G,
        scheduler_D,
        best_psnr,
        best_ssim_at_best_psnr,
        best_epoch,
        train_loader_generator
    )

    # evaluate on full test images every 10 epochs
    if epoch % 10 == 0:
        # --- pre-eval full resumable checkpoint (overwritten each eval) --- #
        save_resume_checkpoint(
            pre_eval_resume_path,
            epoch,
            iteration,
            MyEnsembleNet,
            DNet,
            G_optimizer,
            D_optim,
            scheduler_G,
            scheduler_D,
            best_psnr,
            best_ssim_at_best_psnr,
            best_epoch,
            train_loader_generator
        )
        print(f'[Safety] Saved pre-eval resume checkpoint: {pre_eval_resume_path}')

        try:
            avg_psnr, avg_ssim = evaluate_test_metrics(MyEnsembleNet, test_loader, device, pad_multiple=32)
        except Exception as e:
            print(f'[ERROR] Evaluation failed at epoch {epoch}: {e}')
            print(f'[Recovery] Resume from: {pre_eval_resume_path}')
            raise

        is_best = False
        if avg_psnr > best_psnr:
            is_best = True
        elif avg_psnr == best_psnr and avg_ssim > best_ssim_at_best_psnr:
            is_best = True

        if is_best:
            best_psnr = avg_psnr
            best_ssim_at_best_psnr = avg_ssim
            best_epoch = epoch

            torch.save(
                MyEnsembleNet.state_dict(),
                os.path.join(args.model_save_dir, 'best_psnr.pkl')
            )
            torch.save(
                DNet.state_dict(),
                os.path.join(args.denet_save_dir, 'best_psnr_dnet.pkl')
            )

        append_excel_row(
            args.excel_txt,
            epoch,
            avg_psnr,
            avg_ssim,
            is_best
        )

        print(f'[Test @ epoch {epoch}] PSNR: {avg_psnr:.6f}, SSIM: {avg_ssim:.6f}')
        print(f'[Best so far] epoch: {best_epoch}, best_PSNR: {best_psnr:.6f}, SSIM_at_best_PSNR: {best_ssim_at_best_psnr:.6f}')

    # regular checkpoint save every 500 epochs
    if epoch % 500 == 0:
        torch.save(MyEnsembleNet.state_dict(), os.path.join(args.model_save_dir, f"epoch{epoch}.pkl"))
        torch.save(DNet.state_dict(), os.path.join(args.denet_save_dir, f"epoch{epoch}_dnet.pkl"))